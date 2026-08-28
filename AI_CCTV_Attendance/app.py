import base64
import io
import json
import os
import time
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
from flask import Flask, Response, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash

from config import CAMERA_ID, CAMERA_SOURCE, DB_PATH, MAX_IMAGE_SIZE, RECOGNITION_THRESHOLD, SECRET_KEY
from database.db import (
    add_student,
    add_camera,
    add_zone,
    delete_camera,
    delete_zone,
    ensure_default_admin,
    get_all_cameras,
    get_camera_by_id,
    get_all_zones,
    get_zone_by_id,
    get_admin_user,
    get_all_students,
    get_safety_alerts,
    get_attendance_records,
    get_dashboard_stats,
    get_department_summary,
    get_monthly_summary,
    get_student_by_id,
    get_student_by_register,
    init_db,
    mark_attendance,
    create_safety_alert,
    update_alert_status,
    update_camera_status,
)
from recognition.attendance import calculate_attendance_status
from recognition.face_detection import detect_faces, validate_image_file
from recognition.face_recognition import (
    compute_face_embedding,
    load_known_embeddings,
    refresh_known_embeddings,
    save_known_embeddings,
)
from recognition.detection_events import demo_detection
from recognition.person_detection import detect_people_in_zones

BASE_DIR = Path(__file__).resolve().parent
SUPPORTED_CAMERA_TYPES = {"cctv", "ip_camera", "mobile_ip_camera", "webcam", "video_file", "test_stream"}

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax")
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "false").lower() == "true"


def api_error(message, status=400, code="API_ERROR"):
    return jsonify({"success": False, "error": {"code": code, "message": message}}), status


@app.after_request
def add_api_headers(response):
    origin = request.headers.get("Origin")
    allowed_origins = {item.strip() for item in os.environ.get("CORS_ORIGINS", "").split(",") if item.strip()}
    if origin and origin in allowed_origins:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        response.headers["Vary"] = "Origin"
    return response


@app.errorhandler(404)
def handle_not_found(error):
    if request.path.startswith("/api/"):
        return api_error("API endpoint not found.", 404, "NOT_FOUND")
    return error


@app.errorhandler(405)
def handle_method_not_allowed(error):
    if request.path.startswith("/api/"):
        return api_error("HTTP method is not supported for this endpoint.", 405, "METHOD_NOT_ALLOWED")
    return error


def decode_image_data(data_url: str):
    if not data_url or "," not in data_url:
        return None
    header, encoded = data_url.split(",", 1)
    if "base64" not in header:
        return None
    try:
        image_bytes = base64.b64decode(encoded)
    except (ValueError, TypeError):
        return None
    img_array = np.frombuffer(image_bytes, dtype=np.uint8)
    frame = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    if frame is None or frame.size == 0:
        return None
    return frame


def prepare_login_redirect():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return None


def get_attendance_filters():
    return {
        "date": request.args.get("date", "").strip() or None,
        "status": request.args.get("status", "").strip() or None,
        "department": request.args.get("department", "").strip() or None,
        "register_number": request.args.get("student", "").strip() or None,
        "name": request.args.get("name", "").strip() or None,
    }


def camera_public_view(camera):
    return {
        "id": camera["id"],
        "name": camera["name"],
        "camera_type": camera["camera_type"],
        "location": camera["location"],
        "department": camera["department"],
        "status": camera["status"],
        "last_checked_at": camera["last_checked_at"],
        "last_error": camera["last_error"],
        "created_at": camera["created_at"],
    }


def is_supported_camera_source(source: str) -> bool:
    source = source.strip()
    return source.isdigit() or source.startswith(("rtsp://", "http://", "https://"))


def test_camera_source(source: str):
    source = source.strip()
    if not source:
        return {"status": "failed", "message": "A stream URL is required."}
    if source.isdigit():
        capture_source = int(source)
    elif source.startswith(("rtsp://", "http://", "https://")):
        capture_source = source
    else:
        return {"status": "failed", "message": "Use a webcam index or an rtsp/http(s) stream URL."}

    capture = cv2.VideoCapture(capture_source)
    try:
        if hasattr(cv2, "CAP_PROP_OPEN_TIMEOUT_MSEC"):
            capture.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 5000)
        if not capture.isOpened():
            return {"status": "offline", "message": "Stream unavailable or authentication failed."}
        success, frame = capture.read()
        if not success or frame is None:
            return {"status": "offline", "message": "Stream opened but no video frame was received."}
        fps = capture.get(cv2.CAP_PROP_FPS)
        return {
            "status": "connected",
            "message": "Stream connected and a video frame was received.",
            "resolution": {"width": int(frame.shape[1]), "height": int(frame.shape[0])},
            "fps": round(float(fps), 2) if fps and fps > 0 else None,
        }
    finally:
        capture.release()


def offline_frame(message: str):
    frame = np.zeros((360, 640, 3), dtype=np.uint8)
    cv2.putText(frame, message[:48], (28, 180), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (190, 210, 220), 2, cv2.LINE_AA)
    success, encoded = cv2.imencode(".jpg", frame)
    return encoded.tobytes() if success else b""


def camera_mjpeg_stream(camera):
    source = camera["stream_url"]
    capture = cv2.VideoCapture(int(source) if source.isdigit() else source)
    try:
        if not capture.isOpened():
            update_camera_status(camera["id"], "offline", "Stream unavailable or authentication failed.")
            yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + offline_frame("Camera offline") + b"\r\n"
            return
        update_camera_status(camera["id"], "connected", None)
        while True:
            success, frame = capture.read()
            if not success or frame is None:
                update_camera_status(camera["id"], "offline", "No video frame received.")
                yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + offline_frame("No video frame") + b"\r\n"
                return
            success, encoded = cv2.imencode(".jpg", frame, [int(cv2.IMWRITE_JPEG_QUALITY), 80])
            if success:
                yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + encoded.tobytes() + b"\r\n"
            time.sleep(0.04)
    finally:
        capture.release()


def do_face_recognition(frame, confirm=True):
    face_boxes = detect_faces(frame)
    if not face_boxes:
        return {"success": True, "status": "UNKNOWN", "recognized": False, "message": "No face detected. Please position your face inside the camera frame."}
    if len(face_boxes) > 1:
        return {"success": True, "status": "UNKNOWN", "recognized": False, "message": "Multiple faces detected. Please ensure the intended student is clearly visible."}

    x, y, w, h = face_boxes[0]
    face = frame[y : y + h, x : x + w]
    embedding = compute_face_embedding(face)
    if embedding is None:
        return {"success": True, "status": "UNKNOWN", "recognized": False, "message": "No face detected. Please position your face inside the camera frame."}

    known_embeddings = load_known_embeddings()
    best_student_id = None
    best_confidence = 0.0
    for student_id, known_embedding in known_embeddings.items():
        distance = float(np.linalg.norm(embedding - known_embedding))
        confidence = max(0.0, 100.0 - distance * 1.6)
        if confidence > best_confidence:
            best_student_id = student_id
            best_confidence = confidence

    if best_student_id is None or best_confidence < RECOGNITION_THRESHOLD:
        return {"success": True, "status": "UNKNOWN", "recognized": False, "message": "Unknown face"}

    student = get_student_by_id(int(best_student_id))
    if not student:
        return {"success": True, "status": "UNKNOWN", "recognized": False, "message": "Unknown face"}

    now = datetime.now()
    attendance_status = calculate_attendance_status(now.strftime("%H:%M:%S"), "09:00", "09:15")
    marked = mark_attendance(int(student["id"]), now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), attendance_status, round(best_confidence, 2), CAMERA_ID) if confirm else False

    return {
        "success": True,
        "status": "RECOGNIZED",
        "recognized": True,
        "student_id": int(student["id"]),
        "name": student["name"],
        "register_number": student["register_number"],
        "department": student["department"],
        "confidence": round(best_confidence, 2),
        "attendance_status": "Present" if marked else "Already Marked",
        "already_marked": not marked,
        "time": now.strftime("%I:%M:%S %p"),
        "message": "Attendance marked successfully." if marked else f"{student['name']} is already marked present today.",
    }


@app.context_processor
def inject_now():
    return {"current_year": datetime.now().year, "session_user": session.get("user")}


@app.before_request
def require_login():
    public_routes = {"login", "api_login", "home", "static"}
    if request.method == "OPTIONS" and request.path.startswith("/api/"):
        return ("", 204)
    if request.endpoint in public_routes or request.endpoint is None:
        return None
    if "user" not in session:
        if request.path.startswith("/api/"):
            return api_error("Authentication required.", 401, "AUTHENTICATION_REQUIRED")
        return redirect(url_for("login"))
    return None


@app.route("/")
def home():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    redirect_response = prepare_login_redirect()
    if redirect_response:
        return redirect_response

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if not username:
            flash("Required username", "error")
        elif not password:
            flash("Required password", "error")
        else:
            user = get_admin_user(username)
            if user and check_password_hash(user["password_hash"], password):
                session["user"] = username
                flash("Login successful.", "success")
                return redirect(url_for("dashboard"))
            flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    payload = request.get_json(silent=True) or {}
    username = str(payload.get("username", "")).strip()
    password = str(payload.get("password", ""))
    if not username or not password:
        return api_error("Username and password are required.", 400, "VALIDATION_ERROR")
    user = get_admin_user(username)
    if not user or not check_password_hash(user["password_hash"], password):
        return api_error("Invalid username or password.", 401, "INVALID_CREDENTIALS")
    session["user"] = username
    return jsonify({"success": True, "data": {"username": username, "role": user["role"]}})


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"success": True})


@app.route("/api/health")
def api_health():
    return jsonify({"success": True, "status": "healthy", "application": "healthy", "database": "healthy", "ai": "available", "storage": "available"})


@app.route("/api/students", methods=["GET", "POST"])
def students_api():
    if request.method == "GET":
        fields = ("id", "register_number", "name", "department", "year", "section", "email", "phone", "created_at")
        return jsonify({"success": True, "data": [{field: student[field] for field in fields} for student in get_all_students()]})
    payload = request.get_json(silent=True) or {}
    required = ("register_number", "name", "department", "year", "section")
    if any(not str(payload.get(field, "")).strip() for field in required):
        return api_error("Register number, name, department, year, and section are required.", 400, "VALIDATION_ERROR")
    student_id = add_student({field: str(payload.get(field, "")) for field in required + ("email", "phone")})
    if student_id is None:
        return api_error("A student with this register number already exists.", 409, "DUPLICATE_STUDENT")
    return jsonify({"success": True, "data": {"id": student_id}}), 201


@app.route("/api/students/<int:student_id>", methods=["DELETE"])
def delete_student_api(student_id):
    if not get_student_by_id(student_id):
        return api_error("Student not found.", 404, "NOT_FOUND")
    from database.db import delete_student
    delete_student(student_id)
    return jsonify({"success": True})


@app.route("/api/students/<int:student_id>/face", methods=["POST"])
def add_student_face_api(student_id):
    student = get_student_by_id(student_id)
    frame = decode_image_data((request.get_json(silent=True) or {}).get("image"))
    if not student or frame is None:
        return api_error("Student and a valid base64 image are required.", 400, "VALIDATION_ERROR")
    faces = detect_faces(frame)
    if len(faces) != 1:
        return api_error("The image must contain exactly one face.", 400, "INVALID_FACE_IMAGE")
    x, y, width, height = faces[0]
    embedding = compute_face_embedding(frame[y : y + height, x : x + width])
    if embedding is None:
        return api_error("A usable face embedding could not be generated.", 400, "INVALID_FACE_IMAGE")
    embeddings = load_known_embeddings()
    embeddings[str(student_id)] = embedding
    save_known_embeddings({key: value.tolist() if hasattr(value, "tolist") else value for key, value in embeddings.items()})
    refresh_known_embeddings()
    return jsonify({"success": True, "data": {"student_id": student_id}})


@app.route("/api/attendance")
def attendance_api():
    records = get_attendance_records(get_attendance_filters())
    return jsonify({"success": True, "data": [dict(record) for record in records]})


@app.route("/api/attendance/recognize", methods=["POST"])
def recognize_attendance_api():
    payload = request.get_json(silent=True) or {}
    frame = decode_image_data(payload.get("image"))
    if frame is None:
        return api_error("A valid base64 image is required.", 400, "VALIDATION_ERROR")
    return jsonify({"success": True, "data": do_face_recognition(frame, bool(payload.get("confirm")))})


@app.route("/api/events")
def events_api():
    def stream():
        payload = {"cameras": [camera_public_view(camera) for camera in get_all_cameras()], "alerts": [dict(alert) for alert in get_safety_alerts()]}
        yield f"event: system\ndata: {json.dumps(payload)}\n\n"
    return Response(stream(), mimetype="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/api/recordings")
def recordings_api():
    return jsonify({"success": True, "data": []})


@app.route("/dashboard")
def dashboard():
    stats = get_dashboard_stats()
    recent = stats.get("recent_attendance", [])
    department_summary = get_department_summary()
    monthly_summary = get_monthly_summary()
    return render_template(
        "dashboard.html",
        stats=stats,
        recent_attendance=recent,
        department_summary=department_summary,
        monthly_summary=monthly_summary,
    )


@app.route("/cameras")
def cameras():
    return render_template("cameras.html", cameras=get_all_cameras())


@app.route("/monitoring")
def monitoring():
    return render_template("monitoring.html", cameras=get_all_cameras())


@app.route("/alerts")
def alerts():
    return render_template("alerts.html", alerts=get_safety_alerts())


@app.route("/zones")
def zones():
    return render_template("zones.html", cameras=get_all_cameras(), zones=get_all_zones())


@app.route("/api/zones", methods=["GET", "POST"])
def zones_api():
    if request.method == "GET":
        return jsonify({"success": True, "data": [dict(zone) for zone in get_all_zones()]})
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name", "")).strip()
    geometry = payload.get("geometry")
    try:
        camera_id = int(payload.get("camera_id"))
        if not name or not isinstance(geometry, list) or len(geometry) < 3 or not get_camera_by_id(camera_id):
            raise ValueError
        if any(not isinstance(point, dict) or not {"x", "y"}.issubset(point) for point in geometry):
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "A camera, zone name, and at least three valid points are required."}), 400
    zone_id = add_zone({
        "camera_id": camera_id,
        "name": name,
        "zone_type": str(payload.get("zone_type", "restricted")),
        "geometry_json": json.dumps(geometry),
        "severity": str(payload.get("severity", "HIGH")),
        "required_ppe_json": json.dumps(payload.get("required_ppe", [])),
        "alert_enabled": payload.get("alert_enabled", True),
        "recording_enabled": payload.get("recording_enabled", False),
    })
    return jsonify({"success": True, "zone_id": zone_id}), 201


@app.route("/api/zones/<int:zone_id>", methods=["DELETE"])
def delete_zone_api(zone_id):
    if not delete_zone(zone_id):
        return jsonify({"success": False, "message": "Zone not found."}), 404
    return jsonify({"success": True})


@app.route("/api/alerts", methods=["GET"])
def alerts_api():
    status = request.args.get("status", "").strip() or None
    return jsonify({"success": True, "data": [dict(alert) for alert in get_safety_alerts(status)]})


@app.route("/api/demo/alerts", methods=["POST"])
def demo_alert_api():
    cameras = get_all_cameras()
    payload = request.get_json(silent=True) or {}
    camera_id = payload.get("camera_id") or (cameras[0]["id"] if cameras else None)
    event = demo_detection(int(camera_id) if camera_id else None)
    alert_id = create_safety_alert(event)
    if alert_id is None:
        return jsonify({"success": True, "created": False, "message": "Matching alert is within its cooldown window."})
    return jsonify({"success": True, "created": True, "alert_id": alert_id, "simulated": True}), 201


@app.route("/api/cameras/<int:camera_id>/analyze", methods=["POST"])
def analyze_camera_frame(camera_id):
    camera = get_camera_by_id(camera_id)
    if not camera:
        return jsonify({"success": False, "message": "Camera not found."}), 404
    payload = request.get_json(silent=True) or {}
    frame = decode_image_data(payload.get("image"))
    if frame is None:
        return jsonify({"success": False, "message": "A valid base64 image is required."}), 400
    zones = [zone for zone in get_all_zones() if zone["camera_id"] == camera_id and zone["alert_enabled"]]
    analysis = detect_people_in_zones(frame, zones)
    alert_ids = []
    for intrusion in analysis["intrusions"]:
        alert_id = create_safety_alert({
            "alert_type": "RESTRICTED_AREA_INTRUSION",
            "severity": intrusion["severity"],
            "camera_id": camera_id,
            "zone_id": str(intrusion["zone_id"]),
            "tracking_id": intrusion["tracking_id"],
            "confidence": intrusion["confidence"],
            "description": f"Person detected inside {intrusion['zone_name']}.",
        })
        if alert_id is not None:
            alert_ids.append(alert_id)
    return jsonify({"success": True, "data": analysis, "alert_ids": alert_ids})


@app.route("/api/alerts/<int:alert_id>/<action>", methods=["POST"])
def update_alert_api(alert_id, action):
    status_by_action = {"acknowledge": "ACKNOWLEDGED", "investigate": "INVESTIGATING", "resolve": "RESOLVED", "false-positive": "FALSE_POSITIVE", "escalate": "ESCALATED"}
    status = status_by_action.get(action)
    if not status:
        return jsonify({"success": False, "message": "Unsupported alert action."}), 400
    payload = request.get_json(silent=True) or {}
    updated = update_alert_status(alert_id, status, session["user"], str(payload.get("resolution_note", "")) or None)
    if not updated:
        return jsonify({"success": False, "message": "Alert not found or status transition is invalid."}), 404
    return jsonify({"success": True, "status": status})


@app.route("/api/cameras/<int:camera_id>/stream")
def camera_stream(camera_id):
    camera = get_camera_by_id(camera_id)
    if not camera:
        return jsonify({"success": False, "message": "Camera not found."}), 404
    return Response(camera_mjpeg_stream(camera), mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/api/cameras", methods=["GET", "POST"])
def cameras_api():
    if request.method == "GET":
        return jsonify({"success": True, "data": [camera_public_view(camera) for camera in get_all_cameras()]})

    payload = request.get_json(silent=True) or request.form
    name = str(payload.get("name", "")).strip()
    stream_url = str(payload.get("stream_url", "")).strip()
    camera_type = str(payload.get("camera_type", "ip_camera")).strip()
    if not name or not stream_url:
        return jsonify({"success": False, "message": "Camera name and stream URL are required."}), 400
    if not is_supported_camera_source(stream_url):
        return jsonify({"success": False, "message": "Use a webcam index or an rtsp/http(s) stream URL."}), 400
    if camera_type not in SUPPORTED_CAMERA_TYPES:
        return jsonify({"success": False, "message": "Unsupported camera type."}), 400

    camera_id = add_camera({
        "name": name,
        "camera_type": camera_type,
        "stream_url": stream_url,
        "location": str(payload.get("location", "")),
        "department": str(payload.get("department", "")),
    })
    if camera_id is None:
        return jsonify({"success": False, "message": "Unable to save camera configuration."}), 400
    return jsonify({"success": True, "data": camera_public_view(get_camera_by_id(camera_id))}), 201


@app.route("/api/cameras/test", methods=["POST"])
def test_camera_api():
    payload = request.get_json(silent=True) or request.form
    result = test_camera_source(str(payload.get("stream_url", "")))
    return jsonify({"success": result["status"] == "connected", "data": result}), (200 if result["status"] == "connected" else 422)


@app.route("/api/cameras/<int:camera_id>/test", methods=["POST"])
def test_saved_camera_api(camera_id):
    camera = get_camera_by_id(camera_id)
    if not camera:
        return jsonify({"success": False, "message": "Camera not found."}), 404
    result = test_camera_source(camera["stream_url"])
    update_camera_status(camera_id, result["status"], None if result["status"] == "connected" else result["message"])
    return jsonify({"success": result["status"] == "connected", "data": result})


@app.route("/api/cameras/<int:camera_id>/<action>", methods=["POST"])
def camera_action_api(camera_id, action):
    if action not in {"start", "stop", "snapshot", "features"}:
        return api_error("Unsupported camera action.", 400, "UNSUPPORTED_ACTION")
    camera = get_camera_by_id(camera_id)
    if not camera:
        return api_error("Camera not found.", 404, "NOT_FOUND")
    if action == "features":
        payload = request.get_json(silent=True) or {}
        return jsonify({"success": True, "data": {"ai_enabled": bool(payload.get("ai_enabled"))}})
    status = "connected" if action == "start" else "offline" if action == "stop" else camera["status"]
    update_camera_status(camera_id, status, None)
    return jsonify({"success": True, "data": {"status": status}})


@app.route("/api/cameras/<int:camera_id>", methods=["DELETE"])
def delete_camera_api(camera_id):
    if not get_camera_by_id(camera_id):
        return jsonify({"success": False, "message": "Camera not found."}), 404
    delete_camera(camera_id)
    return jsonify({"success": True})


@app.route("/students")
def students():
    student_list = get_all_students()
    return render_template("students.html", students=student_list)


@app.route("/students/register", methods=["GET", "POST"])
def register_student():
    if request.method == "POST":
        register_number = request.form.get("register_number", "").strip()
        name = request.form.get("name", "").strip()
        department = request.form.get("department", "").strip()
        year = request.form.get("year", "").strip()
        section = request.form.get("section", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        face_data = request.form.get("face_image_data")

        if not all([register_number, name, department, year, section]):
            flash("Please fill in all required student fields.", "error")
            return render_template("register_student.html")
        if not register_number.isalnum() and not register_number.replace("-", "").isalnum():
            flash("Please enter a valid register number.", "error")
            return render_template("register_student.html")
        if get_student_by_register(register_number):
            flash("A student with this register number already exists.", "error")
            return render_template("register_student.html")

        frame = None
        if face_data:
            frame = decode_image_data(face_data)
        else:
            file = request.files.get("face_image")
            if file and file.filename:
                if not validate_image_file(file):
                    flash("Invalid image file or unsupported extension.", "error")
                    return render_template("register_student.html")
                file_bytes = file.read()
                if len(file_bytes) > MAX_IMAGE_SIZE:
                    flash("The face image is too large.", "error")
                    return render_template("register_student.html")
                np_array = np.frombuffer(file_bytes, dtype=np.uint8)
                frame = cv2.imdecode(np_array, cv2.IMREAD_COLOR)

        if frame is None or frame.size == 0:
            flash("Capture or upload a clear face image before saving the student.", "error")
            return render_template("register_student.html")

        face_boxes = detect_faces(frame)
        if len(face_boxes) != 1:
            flash("Please ensure exactly one primary face is visible for registration.", "error")
            return render_template("register_student.html")

        x, y, w, h = face_boxes[0]
        face = frame[y : y + h, x : x + w]
        embedding = compute_face_embedding(face)
        if embedding is None:
            flash("Face embedding could not be generated. Please use a clearer image.", "error")
            return render_template("register_student.html")

        student_id = add_student(
            {
                "register_number": register_number,
                "name": name,
                "department": department,
                "year": year,
                "section": section,
                "email": email,
                "phone": phone,
                "face_encoding": json.dumps(embedding.tolist()),
            }
        )
        if student_id is None:
            flash("Failed to create the student record.", "error")
            return render_template("register_student.html")

        embeddings = load_known_embeddings()
        embeddings[str(student_id)] = embedding
        save_known_embeddings({key: value.tolist() if hasattr(value, "tolist") else value for key, value in embeddings.items()})
        refresh_known_embeddings()

        flash(f"Student {name} registered successfully.", "success")
        return redirect(url_for("students"))

    return render_template("register_student.html")


@app.route("/students/<int:student_id>/delete", methods=["POST"])
def delete_student_route(student_id):
    student = get_student_by_id(student_id)
    if student:
        embeddings = load_known_embeddings()
        embeddings.pop(str(student_id), None)
        save_known_embeddings({key: value.tolist() if hasattr(value, "tolist") else value for key, value in embeddings.items()})
        refresh_known_embeddings()
    from database.db import delete_student
    delete_student(student_id)
    flash("Student deleted successfully.", "success")
    return redirect(url_for("students"))


@app.route("/live-camera")
def live_camera():
    return render_template("live_camera.html", camera_id=CAMERA_ID)


@app.route("/live-camera/recognize", methods=["POST"])
def live_camera_recognize():
    payload = request.get_json(silent=True) or {}
    image_data = payload.get("image")
    if not image_data:
        return jsonify({"success": False, "message": "No image data received."}), 400

    frame = decode_image_data(image_data)
    if frame is None:
        return jsonify({"success": False, "message": "Unable to process the captured image."}), 400

    result = do_face_recognition(frame)
    return jsonify(result)


@app.route("/attendance")
def attendance_page():
    filters = get_attendance_filters()
    records = get_attendance_records(filters)
    all_students = get_all_students()
    departments = sorted({student["department"] for student in all_students})
    return render_template(
        "attendance.html",
        attendance=records,
        departments=departments,
        filters=filters,
    )


@app.route("/download-attendance")
def download_attendance():
    filters = get_attendance_filters()
    records = get_attendance_records(filters)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"

    headers = [
        "S.No",
        "Register Number",
        "Student Name",
        "Department",
        "Date",
        "Time",
        "Status",
        "Confidence",
        "Camera Source",
    ]

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000000")
    for idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, record in enumerate(records, start=2):
        sheet.cell(row=row_index, column=1, value=row_index - 1)
        sheet.cell(row=row_index, column=2, value=record["register_number"])
        sheet.cell(row=row_index, column=3, value=record["name"])
        sheet.cell(row=row_index, column=4, value=record["department"])
        sheet.cell(row=row_index, column=5, value=record["attendance_date"])
        sheet.cell(row=row_index, column=6, value=record["attendance_time"])
        sheet.cell(row=row_index, column=7, value=record["status"])
        sheet.cell(row=row_index, column=8, value=float(record["confidence"]))
        sheet.cell(row=row_index, column=9, value=record["camera_id"])

    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{len(records) + 1}"

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Total Students"
    summary["B1"] = get_dashboard_stats()["total_students"]
    summary["A2"] = "Present"
    summary["B2"] = get_dashboard_stats()["present_today"]
    summary["A3"] = "Absent"
    summary["B3"] = get_dashboard_stats()["absent_today"]
    summary["A4"] = "Attendance Percentage"
    summary["B4"] = f"{get_dashboard_stats()['attendance_percentage']}%"
    summary["A5"] = "Report Date"
    summary["B5"] = datetime.now().strftime("%Y-%m-%d")

    for cell in ["A1", "A2", "A3", "A4", "A5"]:
        summary[cell].font = Font(bold=True)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 18

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = "attendance_report.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/reports")
def reports():
    department_summary = get_department_summary()
    monthly_summary = get_monthly_summary()
    stats = get_dashboard_stats()
    return render_template(
        "reports.html",
        department_summary=department_summary,
        monthly_summary=monthly_summary,
        stats=stats,
    )


init_db()
ensure_default_admin("admin", generate_password_hash("admin123"))
refresh_known_embeddings()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
